import os
import openpyxl
from datetime import datetime
from dotenv import load_dotenv
import pandas as pd

#função validador_horario verifica o horario da primeira linha da coluna na planilha base.
#caso esteja atrasado (dia anterior) apresenta um erro e pede para que a planilha base seja atualizada
#caso contrario, segue executando o codigo.

load_dotenv()
def validador_horario(caminho_arquivo, nome_aba, coluna):
    
    workbook = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    sheet = workbook['Consolidado']
    data_planilha = sheet.cell(row=2, column=21).value #primeira linha da coluna da data Atualizacao_planilha
    
    data_hoje = datetime.now().date()
    data_planilha = datetime.strptime(data_planilha, '%d/%m/%Y %H:%M:%S').date()
    
    if( data_planilha == data_hoje):
        return True
    else:
        return False
    
    
#     try: 
#         workbook = openpyxl.load_workbook(caminho_arquivo, data_only=True)
#         sheet = workbook[nome_aba]
#         data_atual = datetime.now().date()
#         datas_invalidas = []

#         for row in range(2, 3):
#             try:
#                 celula = sheet[f"{coluna}{row}"]
#                 valor_celula = datetime.strptime(celula.value, '%d/%m/%Y %H:%M:%S').date()# celula.value.date()
#                 print(type(valor_celula))

#                 if valor_celula is None:
#                     continue  
#                 if isinstance(valor_celula, datetime):
#                     data_celula = valor_celula.date()
#                 elif isinstance(valor_celula, str):
                    
#                     try:
#                         data_celula = datetime.strptime(celula.value, '%d/%m/%Y %H:%M:%S').date()
#                     except ValueError:
#                         try:
#                             data_celula = datetime.strptime(celula.value, '%d/%m/%Y %H:%M:%S').date()
#                         except ValueError:
#                             raise ValueError("Formato desconhecido")
#                 else:
#                     raise TypeError("Tipo de dado não reconhecido")

#                 if data_celula != data_atual:
#                     datas_invalidas.append((row, data_celula))

#             except (ValueError, TypeError):
#                 print(type(valor_celula), type(data_atual))
#                 datas_invalidas.append((row, "Formato inválido"))

#         return datas_invalidas

#     except FileNotFoundError:
#         print(f"Erro: arquivo não encontrado em {caminho_arquivo}")
#         return None
#     except KeyError:
#         print(f"Erro: Aba '{nome_aba}' não encontrada no arquivo.")
#         return None
#     except Exception as e:
#         print(f"Ocorreu um erro inesperado: {e}")
#         return None

# caminho = os.getenv('caminho_validador')
# aba = "Consolidado"
# coluna = "U"

# datas_invalidas = validador_horario(caminho, aba, coluna)

# if datas_invalidas is not None:
#     if datas_invalidas:
#         erros = []
#         print("Primeira linha com data atrasada encontrada:")
#         linha, data = datas_invalidas[0]
#         print(f"Linha {linha}: {data}")
#         msg = "Erro: planilha BASE não está atualizada!!!!!!"
#         print(msg)
#         erros.append({
#             "Pasta": "BASE",
#             "Erro": msg,
#             "DataHora": datetime.now().strftime("%d/%m/%Y %H:%M:%S")
#         })
#         if erros:
#             df_erros = pd.DataFrame(erros)
#             log_path = os.path.join ("log_geral.xlsx")
#             df_erros.to_excel(log_path, index=False)
#             print(f"\nLog de erros da planilha geral salvo em: {log_path}")
#         else:
#             print("\n A pasta geral foi atualizada com sucesso!")
#         exit()
    

